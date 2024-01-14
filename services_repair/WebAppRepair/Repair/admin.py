from django.contrib import admin
from django.http import HttpResponse
from openpyxl.reader.excel import load_workbook

#from django_object_actions import DjangoObjectActions

# Register your models here.

from .models import *

admin.site.register([Client,
                     Employee,
                     Product_Type,
                     Type_of_pockets,
                     Size,
                     Color,
                     Type_of_fastener,
                     Measure,
                     Status,
                     Unit_of_measurement,
                     Service,
                     Stock,
                     Product,
                     CustomUser,
                     Map_of_measurements,
                     Order_status,

                     ])

class Price_list_services_pozitionInline(admin.TabularInline):
    model = Price_list_services_pozition


class Price_list_servicesAdmin(admin.ModelAdmin):
    inlines = [Price_list_services_pozitionInline]
    # search_fields = ("title__startswith",)


admin.site.register(Price_list_services, Price_list_servicesAdmin)

class Price_list_material_pozitionInline(admin.TabularInline):
    model = Price_list_material_pozition


class Price_list_materialAdmin(admin.ModelAdmin):
    inlines = [Price_list_material_pozitionInline]
    # search_fields = ("title__startswith",)


admin.site.register(Price_list_material, Price_list_materialAdmin)

class Materials_orderInline(admin.TabularInline):
    model = Materials_order

class Map_of_measurementsInline(admin.TabularInline):
    model = Map_of_measurements

class Order_pozitionInline(admin.TabularInline):
    model = Order_pozition

class Order_statusInline(admin.TabularInline):
    model = Order_status

class Material_characteristicsInline(admin.TabularInline):
    model = Material_characteristics

class MaterialAdmin(admin.ModelAdmin):
    inlines = [Material_characteristicsInline]

admin.site.register(Material, MaterialAdmin)

class Materials_accounting_journal_pozitionInline(admin.TabularInline):
    model = Materials_accounting_journal_pozition

class Materials_accounting_journalAdmin(admin.ModelAdmin):
    inlines = [Materials_accounting_journal_pozitionInline]

admin.site.register(Materials_accounting_journal, Materials_accounting_journalAdmin)

class OrderAdmin(admin.ModelAdmin):
    inlines = ([Materials_orderInline, Map_of_measurementsInline, Order_pozitionInline, Order_statusInline])
    fieldsets = [
        (None, {'fields': [('number', 'date'), 'client', 'employee', 'product', 'NDS']}),
        ('ПРИМЕРКА', {'fields': [('date_of_the_first_fitting', 'date_of_the_second_fitting'),]}),
        ('ИЗГОТОВЛЕНИЕ', {'fields': [('planned_production_date', 'actual_production_date'), ]}),
        #('Сумма заказа', {'fields': ['total_atelier_materials_cost', 'total_client_materials_cost', ]}),
    ]
    list_filter = ('date', 'client',)

    change_form_template = "admin/change_form.html"  # Here

    actions = ['_generate_a_receipt']

    def _generate_a_receipt(self, request, queryset):
        template = 'templates/receipt_maket.xlsx'
        wb = load_workbook(template)
        ws = wb.active
        field_mapping = {
            'number': 'C1',
            'date': 'E1',
            'client': 'F3',
            'address': 'F4',
            'telefon_number': 'F5',
            'product': 'F6',
            'employee': 'F7',
            'date_1_fitting': 'C23',
            'date_2_fitting': 'C24',
            'fact_date': 'C25'

        }

        clients = Client.objects.all()
        employees = Employee.objects.all()
        products = Product.objects.all()
        orders = Order.objects.all()

        for agreement in queryset:

            for field, cell in field_mapping.items():

                if field == 'client':

                    client = clients.get(id=agreement.client_id)

                    value = client.name if client else ''

                elif field == 'address':

                    client = clients.get(id=agreement.client_id)

                    value = client.adress if client else ''

                elif field == 'telefon_number':

                    client = clients.get(id=agreement.client_id)

                    value = client.number_phone if client else ''

                elif field == 'product':

                    product = products.get(id=agreement.product_id)

                    value = product.name if product else ''

                elif field == 'employee':

                    employee = employees.get(id=agreement.employee_id)

                    value = f"{employee.last_name} {employee.first_name} {employee.middle_name}" if employee else ''

                elif field == 'date_1_fitting':

                    order = orders.get(id=agreement.order_id)

                    value = order.date_of_the_first_fitting if order else ''

                elif field == 'date_2_fitting':

                    order = orders.get(id=agreement.order_id)

                    value = order.date_of_the_second_fitting if order else ''



                else:

                    value = getattr(agreement, field)

                ws[cell] = value



            # Создаем новый файл для сохранения данных

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        response['Content-Disposition'] = 'attachment; filename=Договор проката.xlsx'

        # Сохраняем данные в новый файл

        wb.save(response)

        return response

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        extra_context = extra_context or {}

        extra_context['generate_a_receipt'] = True  # Here

        return super().changeform_view(request, object_id, form_url, extra_context)








    # search_fields = ("title__startswith",)


admin.site.register(Order, OrderAdmin)

admin.site.site_title = 'Главная страница'
admin.site.index_title = 'Главная страница'
admin.site.site_header = 'Ателье по ремонту одежды'
